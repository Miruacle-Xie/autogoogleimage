# -- coding: UTF-8 --
import re
import os
import sys
import time

import pyperclip

import win32gui
import win32con
import win32clipboard

from openpyxl import load_workbook
from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains


def configchrome():
    try:
        chrome_options = Options()
        chrome_options.add_argument('--incognito')
        driver = webdriver.Chrome(chrome_options=chrome_options)
        return driver
    except Exception as e:
        print(repr(e))
        return False


def googleimage(driver, filepathlist):
    resultlist = []
    urllist = []
    resetflag = False
    refreshflag = False
    breakloopflag = False
    elfindflag = False
    cantfindflag = False
    lasturl = ""
    for i in range(len(filepathlist)):
        print("\n{}: {}".format(i+1, filepathlist[i]))
        if i == 0 or resetflag:
            try:
                print("Line:{}, 初始化中...".format(sys._getframe().f_lineno))
                driver.get("https://lens.google.com/search?p=AcLkwR1czMJwA0je-PjQsowwECIoS0DUK0scf\
                    BkBg5ahrkKQUftizfjUwhPu63efjg4pOPNEUW8pTKa6Zc84fC3dm9uPpV5XpQpJqpCIjE0CbpcEHpR\
                    FXjuG1T2VdmZYmuKGO7salXL7TfDHKLTL-Dxh_RykCDdr4Vi12aOyAmsQMWgk4sCIX4nWwBxDbThjr\
                    VgurR24w9J6wuQaSI5PaWlnLXZElp14-A39uvcJU45aep-ea5D3EvAWY2xb1h4U8GrO6XWUJDMuLY3\
                    MUdSk7gSSFP8evWybut1l72YnrqLVYTzf8uqLqIx6heFWLUICCSVhfzhXYhiDqfHGsrMTbBCLyR_Io\
                    CdqymS5Li-fdmACbjdV&ep=gisbubb&hl=en&re=df#lns=W251bGwsbnVsbCxudWxsLG51bGwsbnV\
                    sbCxudWxsLG51bGwsIkVrY0tKREJrTlRCbFpHTTNMVFU0TXpndE5EVmxOeTA1TTJGbUxURmlNakEyT\
                    ldJNFpUWTFNQklmU1RkdFRIWnlaVzQyT1hOWmQwMURZVjlYYUdaVlQxZ3RNVVUzTVZKNFp3PT0iXQ==")
            except Exception as e1:
                print(repr(e1))
                print("第一次异常")
                return resultlist

            try:
                print("Line:{}, 进行Upload操作中...".format(sys._getframe().f_lineno))
                lasturl = driver.current_url
                buttonUpload = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[aria-haspopup="menu"]')))
                buttonUpload.click()
                time.sleep(2)
                WebDriverWait(driver, 30).until(
                    EC.visibility_of_element_located((By.CLASS_NAME, 'google-material-icons')))
                inputimage = driver.find_elements_by_class_name('google-material-icons')
                for inputimage_tmp in inputimage:
                    # print(inputimage_tmp.get_attribute("textContent"))
                    if inputimage_tmp.get_attribute("textContent") == "laptop_chromebook":
                        ActionChains(driver).move_to_element(inputimage_tmp).click(inputimage_tmp).perform()
                        # print("1")
                stime = time.time()
                while True:
                    handle = win32gui.FindWindow("#32770", "打开")
                    # print(handle)
                    if handle:
                        # print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
                        win32gui.PostMessage(handle, win32con.WM_CLOSE, 0, 0)
                        break
                    if time.time() - stime > 10:
                        print("弹窗超时")
                        break
            except Exception as e2:
                print(repr(e2))
                print("upload步骤异常")
                return resultlist

            try:
                print("\nLine:{}, 获取图片输入元素中...".format(sys._getframe().f_lineno))
                inputel = WebDriverWait(driver, 30).until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, 'input[type="file"]')))
                resetflag = False
            except Exception as e3:
                print(repr(e3))
                print("获取图片输入元素异常")
                return resultlist

        try:
            print("Line:{}, 图片输入中...".format(sys._getframe().f_lineno))
            print(inputel)
            # print(filepathlist[i])
            inputel.send_keys(filepathlist[i])
            if i != 0:
                urllist.append(lasturl)
            stime = time.time()
            while lasturl == driver.current_url:
                if time.time() - stime > 30:
                    print("加载图片超时")
                    refreshflag = True
                    break
            time.sleep(1)
            print(lasturl)
            if refreshflag:
                refreshflag = False
                resetflag = True
                resultlist.append("加载图片超时")
                lasturl = driver.current_url
                print(lasturl)
                continue
            # time.sleep(3)
        except Exception as e4:
            resetflag = True
            resultlist.append("ERROR")
            print(repr(e4))
            print("获取图片输入元素异常")
            lasturl = driver.current_url
            urllist.append(lasturl)
            print(lasturl)
            continue

        try:
            print("\nLine:{}, Switch to Text mode...".format(sys._getframe().f_lineno))
            textbutton = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[aria-label="Switch to Text mode"]')))
            # print(textbutton)
            textbutton.click()
            # time.sleep(1)
        except Exception as e5:
            # resetflag = True
            resultlist.append("ERROR")
            print(repr(e5))
            print("切换Text异常")
            lasturl = driver.current_url
            urllist.append(lasturl)
            print(lasturl)
            continue

        try:
            print("\nLine:{}, Select all text...".format(sys._getframe().f_lineno))
            stime = time.time()
            while True:
                cantfind = re.search(r"Can't find text", driver.page_source)
                if cantfind:
                    cantfindflag = True
                    break
                alltextbutton = WebDriverWait(driver, 30).until(
                    EC.visibility_of_any_elements_located((By.TAG_NAME, "button")))
                # print(alltextbutton)
                for alltextbutton_tmp in alltextbutton:
                    # print(alltextbutton_tmp.text)
                    if alltextbutton_tmp.text == "Select all text":
                        alltextbutton_tmp.click()
                        elfindflag = True
                        break
                if elfindflag:
                    elfindflag = False
                    break
                if time.time() - stime > 15:
                    print("未找到Select all text超时")
                    breakloopflag = True
                    break
            if cantfindflag:
                cantfindflag = False
                resultlist.append("无文字")
                print("无文字")
                lasturl = driver.current_url
                print(lasturl)
                continue
            if breakloopflag:
                breakloopflag = False
                resultlist.append("未找到Select all text")
                print("未找到Select all text")
                lasturl = driver.current_url
                print(lasturl)
                continue
        except Exception as e6:
            # resetflag = True
            resultlist.append("ERROR")
            print(repr(e6))
            print("Select all text异常")
            lasturl = driver.current_url
            urllist.append(lasturl)
            print(lasturl)
            continue

        try:
            print("\nLine:{}, Copy text...".format(sys._getframe().f_lineno))
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.CloseClipboard()
            stime = time.time()
            while True:
                copybutton = WebDriverWait(driver, 30).until(
                    EC.visibility_of_any_elements_located((By.TAG_NAME, "button")))
                # print(copybutton)
                for copybutton_tmp in copybutton:
                    # print(copybutton_tmp.text)
                    if copybutton_tmp.text == "Copy text":
                        copybutton_tmp.click()
                        elfindflag = True
                        break
                if elfindflag:
                    elfindflag = False
                    break
                if time.time() - stime > 5:
                    print("未找到Copy text")
                    breakloopflag = True
                    break
            if breakloopflag:
                breakloopflag = False
                resultlist.append("Copy text")
                print("未找到Copy text")
                lasturl = driver.current_url
                print(lasturl)
                continue
        except Exception as e7:
            # resetflag = True
            resultlist.append("ERROR")
            print(repr(e7))
            print("Copy text异常")
            lasturl = driver.current_url
            urllist.append(lasturl)
            print(lasturl)
            continue

        try:
            print("\nLine:{}, 获取粘贴板信息...".format(sys._getframe().f_lineno))
            ocrtext = pyperclip.paste()
            # win32clipboard.OpenClipboard()
            # ocrtext = win32clipboard.GetClipboardData(win32con.CF_UNICODETEXT)
            # win32clipboard.CloseClipboard()
            resultlist.append(ocrtext)
        except Exception as e8:
            # resetflag = True
            resultlist.append("ERROR")
            print(repr(e8))
            print("获取粘贴板数据异常")
            lasturl = driver.current_url
            urllist.append(lasturl)
            print(lasturl)
            continue
        lasturl = driver.current_url
        print(lasturl)
        time.sleep(3)
    urllist.append(lasturl)
    # win32clipboard.CloseClipboard()
    return resultlist, urllist


def main():
    filepath = input("请输入文件夹路径:\n")
    filepath = filepath.replace("\"", "").replace("\'", "")
    print(os.listdir(filepath))
    filelist = [filepath + "\\" + i for i in os.listdir(filepath) if os.path.splitext(i)[1].lower() == ".jpg" or os.path.splitext(i)[1].lower() == ".png"]
    chromedriver = configchrome()
    time_start = time.time()
    if chromedriver:
        result, url = googleimage(chromedriver, filelist)
        print(result)
        # result = [result_tmp.replace("\r\n", " ") for result_tmp in result]
        # for i in result:
        #     print(i)
        chromedriver.quit()
        # zipresult = [",".join(list(i)) + "\n" for i in zip(filelist, result)]
        # print(zipresult)
        resultreport = filepath + "\\report.xlsx"
        if os.path.exists(resultreport):
            os.remove(resultreport)
        wb = Workbook()
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]  # index为0为第一张表
        for cnt in range(1, len(filelist) + 1):
            try:
                print(filelist[cnt - 1])
                ws.cell(cnt, 1).value = filelist[cnt - 1]
                ws.cell(cnt, 2).value = result[cnt - 1]
                ws.cell(cnt, 3).value = url[cnt - 1]
            except Exception as e:
                print(filelist[cnt - 1] + "写入异常")
                print(repr(e))
        wb.save(resultreport)
        time_end = time.time()
        input("已生成报告, 耗时时间:{}, 平均耗时:{}, 按回车键结束".format(time_end - time_start, (time_end - time_start) / len(filelist)))
    else:
        print("启动Chrome异常")
    return


if __name__ == "__main__":
    main()
    # a = ["F:\\JetBrains\\officeTools\\autogoogle\\pic\\1.jpg", "F:\\JetBrains\\officeTools\\autogoogle\\pic\\4.jpg"]
    # b = ["凍吃\r\nTuch Theat\r\nHappy B Halloween!", "Halloween\r\nRIP"]
    # b = [result_tmp.replace("\r\n", " ") for result_tmp in b]
    #
    # c = [",".join(list(i)) + "\n" for i in zip(a, b)]
    # print(c)
    # # c = ['F:\\JetBrains\\officeTools\\autogoogle\\pic\\1.jpg,凍吃\r\nTuch Theat\r\nHappy B Halloween!\n', 'F:\\JetBrains\\officeTools\\autogoogle\\pic\\4.jpg,Halloween\r\nRIP\n']
    # c = ['F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\1.jpg,凍吃 Tuch Theat Happy B Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\10.jpg,+\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\100.jpg,TRICK TREAT\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\101.jpg,7 Happy HALLOWEEN Party C\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\102.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\103.jpg,HAPPY HALLOWEEN OCT 31\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\104.jpg,HAPPY Halloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\105.jpg,Halloween party\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\106.jpg,Happy HALLOWEEN Party TRICK OR TREAT :( :\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\107.jpg,HAPPY HALLOWEEN TRICK OR TREAT\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\108.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\109.jpg,Halloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\11.jpg,Halloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\110.jpg,WEEN HALLOW HALLOWEEN HELLOWEE Alloween HELLOWEEN 海味\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\111.jpg,WEEN HALLOW HALLOWEEN HELLOWEE Alloween HELLOWEEN 海味\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\112.jpg,8 de op m 8 80\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\113.jpg,EDIEDI DEDI EDI DDDD H\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\114.jpg,НАРРУ. 3 HALLOWEEN 35\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\115.jpg,HAPPY HALLOWEEN TRICK OR TREAT\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\116.jpg,RIP ww RIP HAPPY SPRING RIP www\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\117.jpg,B% TREAT TRICK RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\118.jpg,B% TREAT TRICK RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\119.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\12.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\120.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\121.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\122.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\123.jpg,Trick ՉԻ Treat\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\124.jpg,Hloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\125.jpg,Happy\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\126.jpg,Hallowen allowen Happy\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\127.jpg,8 TRICK TREAT! OR HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\128.jpg,HELLO! HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\129.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\12_B095BVKKR8_Halloween Boo Doormat Floor Door Rug Outdoor &Indoor Gnome Pumpkin Door Mats for Home Entrance Kitchen Bathroom Living Room 15.7 x 23.6 Inch.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\13.jpg,3\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\130.jpg,Happy\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\131.jpg,BOO! BOO! BOO!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\132.jpg,+ * RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\133.jpg,+ * RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\134.jpg,+ * RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\135.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\136.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\137.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\138.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\139.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\14.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\140.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\141.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\142.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\143.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\144.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\145.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\146.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\147.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\148.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\149.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\15.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\150.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\151.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\152.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\153.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\154.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\155.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\155235_wallpapers-halloween-wallpapers_1600x1200_h.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\156.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\157.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\158.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\159.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\16.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\160.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\1602082361.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\161.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\162.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\163.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\164.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\165.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\166.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\167.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\168.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\169.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\17.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\170.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\171.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\172.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\173.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\174.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\175.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\176.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\177.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\178.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\179.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\18.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\180.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\181.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\182.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\183.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\184.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\185.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\186.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\187.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\188.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\189.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\19.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\190.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\191.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\192.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\193.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\194.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\195.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\196.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\197.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\198.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\199.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\1_B095B9P3JN_Halloween Boo Doormat Floor Door Rug Outdoor &Indoor Gnome Pumpkin Door Mats for Home Entrance Kitchen Bathroom Living Room 15.7 x 23.6 Inch.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\20.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\200.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\201.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\202.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\203.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\204.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\205.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\206.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\207.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\208.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\209.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\21.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\210.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\211.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\212.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\213.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\214.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\215.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\216.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\217.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\218.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\219.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\22.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\220.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\221.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\222.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\223.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\224.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\225.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\226.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\227.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\228.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\229.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\23.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\230.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\231.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\232.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\233.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\234.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\235.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\236.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\237.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\238.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\239.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\24.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\240.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\241.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\242.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\243.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\244.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\245.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\246.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\247.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\248.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\249.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\25.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\250.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\251.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\252.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\253.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\254.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\255.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\256.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\257.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\258.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\259.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\26.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\260.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\261.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\262.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\263.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\264.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\265.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\266.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\267.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\268.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\269.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\27.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\270.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\271.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\272.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\273.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\274.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\275.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\276.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\277.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\278.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\279.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\28.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\280.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\281.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\282.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\283.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\284.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\285.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\286.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\287.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\288.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\289.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\29.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\290.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\291.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\292.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\293.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\294.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\295.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\296.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\297.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\298.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\299.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2_B07Y832WQN_Festive Halloween Doormat Orange Black Pumpkin Various Spooks Witch Broom Bat Cat Mouse Spider Skull Door Mats Durable Soft Carpet Non Slip Rubber Backing Rug Nice Mat for Mouse Home Decoration Entry.png,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2_B09BCL9VCN_Halloween Doormats Home Front Door Decorations Welcome Blankets Halloween Festival Decor Door Mat Anti-Slip Back Indoor Outdoor Decor Carpet Entrance Rugs-40x60cm.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2_B09BKMFWFT_Halloween Doormat Blanket Cartoon Floor Mat Halloween Decor Door Mat Soft Anti-Slip Bottom Indoor Outdoor Carpet Hallway Kitchen Absorbent Mat Foot Mat (O-1pcs).jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\3.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\30.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\300.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\301.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\302.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\31.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\32.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\33.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\34.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\35.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\36.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\37.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\38.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\39.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\4.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\40.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\41.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\42.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\43.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\44.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\45.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\46.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\46_B098J2GVC4_Halloween Boo Doormats Rugs- Ghost Spooky Pumpkin Witches Hat Non-Slip Entrance Door Mats Carpet Indoor for HomeBathroomKitchenBedroom, Small 23.6x15.7inch, Buffalo Check Plaid Orange Black.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\47.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\48.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\49.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\5.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\50.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\51.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\52.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\53.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\54.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\55.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\56.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\57.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\58.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\59.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\6.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\60.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\61.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\62.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\63.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\64.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\65.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\66.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\67.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\68.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\69.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\7.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\70.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\71.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\72.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\73.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\74.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\75.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\76.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\77.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\78.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\79.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\8.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\80.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\81.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\82.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\83.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\84.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\85.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\86.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\87.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\88.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\89.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\9.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\90.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\91.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\92.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\93.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\94.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\95.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\96.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\97.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\98.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\99.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\EIOkthVXkAIxZBI.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween door mat-xzx-210825.xlsx,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween-silhouette-set-vector-6109744.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween.jpeg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween.png,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\khLni7.png,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\Thumbs.db,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\vintage-happy-halloween-hanging-decorations-paper-vector-24359728.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\指纹图片名称 halloween-210826.txt,ERROR\n']
    # c = ['F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\1.jpg,凍吃 Tuch Theat Happy B Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\10.jpg,+\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\100.jpg,TRICK TREAT\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\101.jpg,7 Happy HALLOWEEN Party C\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\102.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\103.jpg,HAPPY HALLOWEEN OCT 31\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\104.jpg,HAPPY Halloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\105.jpg,Halloween party\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\106.jpg,Happy HALLOWEEN Party TRICK OR TREAT :( :\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\107.jpg,HAPPY HALLOWEEN TRICK OR TREAT\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\108.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\109.jpg,Halloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\11.jpg,Halloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\110.jpg,WEEN HALLOW HALLOWEEN HELLOWEE Alloween HELLOWEEN 海味\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\111.jpg,WEEN HALLOW HALLOWEEN HELLOWEE Alloween HELLOWEEN 海味\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\112.jpg,8 de op m 8 80\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\113.jpg,EDIEDI DEDI EDI DDDD H\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\114.jpg,НАРРУ. 3 HALLOWEEN 35\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\115.jpg,HAPPY HALLOWEEN TRICK OR TREAT\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\116.jpg,RIP ww RIP HAPPY SPRING RIP www\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\117.jpg,B% TREAT TRICK RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\118.jpg,B% TREAT TRICK RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\119.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\12.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\120.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\121.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\122.jpg,Happy Halloween!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\123.jpg,Trick ՉԻ Treat\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\124.jpg,Hloween\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\125.jpg,Happy\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\126.jpg,Hallowen allowen Happy\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\127.jpg,8 TRICK TREAT! OR HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\128.jpg,HELLO! HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\129.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\12_B095BVKKR8_Halloween Boo Doormat Floor Door Rug Outdoor &Indoor Gnome Pumpkin Door Mats for Home Entrance Kitchen Bathroom Living Room 15.7 x 23.6 Inch.jpg,HAPPY HALLOWEEN\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\13.jpg,3\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\130.jpg,Happy\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\131.jpg,BOO! BOO! BOO!\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\132.jpg,+ * RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\133.jpg,+ * RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\134.jpg,+ * RIP\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\135.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\136.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\137.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\138.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\139.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\14.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\140.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\141.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\142.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\143.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\144.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\145.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\146.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\147.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\148.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\149.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\15.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\150.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\151.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\152.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\153.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\154.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\155.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\155235_wallpapers-halloween-wallpapers_1600x1200_h.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\156.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\157.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\158.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\159.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\16.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\160.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\1602082361.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\161.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\162.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\163.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\164.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\165.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\166.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\167.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\168.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\169.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\17.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\170.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\171.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\172.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\173.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\174.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\175.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\176.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\177.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\178.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\179.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\18.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\180.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\181.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\182.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\183.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\184.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\185.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\186.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\187.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\188.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\189.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\19.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\190.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\191.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\192.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\193.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\194.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\195.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\196.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\197.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\198.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\199.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\1_B095B9P3JN_Halloween Boo Doormat Floor Door Rug Outdoor &Indoor Gnome Pumpkin Door Mats for Home Entrance Kitchen Bathroom Living Room 15.7 x 23.6 Inch.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\20.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\200.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\201.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\202.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\203.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\204.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\205.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\206.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\207.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\208.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\209.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\21.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\210.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\211.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\212.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\213.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\214.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\215.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\216.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\217.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\218.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\219.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\22.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\220.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\221.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\222.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\223.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\224.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\225.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\226.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\227.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\228.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\229.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\23.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\230.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\231.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\232.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\233.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\234.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\235.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\236.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\237.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\238.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\239.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\24.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\240.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\241.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\242.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\243.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\244.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\245.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\246.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\247.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\248.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\249.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\25.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\250.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\251.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\252.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\253.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\254.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\255.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\256.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\257.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\258.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\259.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\26.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\260.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\261.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\262.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\263.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\264.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\265.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\266.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\267.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\268.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\269.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\27.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\270.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\271.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\272.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\273.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\274.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\275.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\276.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\277.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\278.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\279.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\28.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\280.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\281.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\282.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\283.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\284.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\285.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\286.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\287.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\288.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\289.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\29.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\290.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\291.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\292.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\293.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\294.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\295.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\296.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\297.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\298.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\299.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2_B07Y832WQN.png,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2_B09BCL9VCN.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\2_B09BKMFWFT.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\3.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\30.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\300.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\301.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\302.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\31.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\32.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\33.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\34.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\35.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\36.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\37.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\38.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\39.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\4.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\40.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\41.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\42.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\43.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\44.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\45.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\46.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\46_B098J2GVC4_.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\47.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\48.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\49.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\5.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\50.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\51.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\52.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\53.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\54.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\55.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\56.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\57.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\58.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\59.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\6.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\60.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\61.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\62.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\63.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\64.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\65.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\66.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\67.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\68.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\69.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\7.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\70.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\71.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\72.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\73.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\74.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\75.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\76.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\77.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\78.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\79.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\8.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\80.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\81.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\82.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\83.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\84.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\85.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\86.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\87.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\88.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\89.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\9.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\90.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\91.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\92.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\93.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\94.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\95.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\96.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\97.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\98.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\99.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\EIOkthVXkAIxZBI.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween door mat-xzx-210825.xlsx,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween-silhouette-set-vector-6109744.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween.jpeg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\halloween.png,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\khLni7.png,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\Thumbs.db,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\vintage-happy-halloween-h.jpg,ERROR\n', 'F:\\file share\\pic\\已上传指纹\\8.25超级团队xzx\\指纹图片名称 halloween-210826.txt,ERROR\n']
    # fp = open("F:\\JetBrains\\officeTools\\autogoogle\\pic\\"+"1.csv", 'a', encoding='utf-8')
    # for i in c:
    #     fp.write(i)
    # fp.close()
    # a = [b'\x83\xf6\xb3\xd4\r\nTuch Theat\r\nHappy B Halloween!', b'Halloween\r\nRIP']
    # resultfile = "F:\\JetBrains\\officeTools\\autogoogle\\pic\\1.xlsx"
    # if os.path.exists(resultfile):
    #     os.remove(resultfile)
    # wb = Workbook()
    # sheetnames = wb.sheetnames
    # ws = wb[sheetnames[0]]  # index为0为第一张表
    # for i in range(1, len(a) + 1):
    #     try:
    #         print(a[i - 1])
    #         ws.cell(i, 1).value = a[i - 1].decode('utf-8')
    #         # ws.cell(i, 2).value = b[i - 1]
    #     except Exception as e:
    #         print(a[i - 1] + "写入异常")
    #         print(repr(e))
    # wb.save(resultfile)
